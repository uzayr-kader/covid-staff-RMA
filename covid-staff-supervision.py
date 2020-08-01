'''
Go through each person’s timetable one by one.
Where they have a non-teaching lesson (identifiable from a set of codes) they need to
    look at the lesson before and see where they were teaching and put this ‘block’ down as their supervision area for first five minutes of the lesson.
    Then they need to look at the lesson after and see where they are due to teach next and put this block down as their supervision for the last five minutes of the lesson. 
    This needs doing again for every member of teaching staff and adding to the supervision tables I have for each block.

The codes to be looked up are:
·         PPA
·         NQT time
·         Mentor
·         Fac. Time/Cover
·         TLR time
·         SLT time
·         NTE
·         ScItt (except if it is on a Friday)
·         HoY duties
·         DUTY (P2 only - not P4)
Also, if the cell is completely empty please treat it as one of the above. If the cell contains BLANK please do not treat it as one of the above.

If a member of staff has any of the above codes on a particular lesson then can the following happen:
·         Look at the room number in the cell to the left and compare to the room codes on each of the tabs of three supervision staffing.
·         Where there is a match copy the three letter staff code into the relevant ‘End of Period…’ slot for the relevant day and lesson
·         If there is no room code in the adjacent cell then go to the tab that says ‘KS4 outside’
·         Look at the room number in the cell to the right and compare to the room codes on each of the tabs of three supervision staffing.
·         Where there is a match copy the three letter staff code into the relevant ‘Start of Period…’ slot for the relevant day and lesson
·         If there is no room code in the adjacent cell then go to the tab that says ‘KS4 outside’

Please note – Wednesday needs to be done for both Week A and Week B as this is the day that has a two week timetable. All other days can be done for week A only. I have put two sets of columns for Wednesdays to allow for this.

The ‘End of form/Start of P1’ and ‘End of P5/Start of form’ columns will have to remain blank for now as I do not have the full information about when each co-tutor is taking the tutor group so this will affect these two times. Once Julie and Sophie confirm that all of the tutor information is in and sends it to me I can enter that manually for each day and block.
'''

#### Import necessary packages ####
from openpyxl import *  # Write to excel
import re               # Regex
from enum import Enum   # Enumeration


#### Class definitions ####
class Days(Enum):
    Monday = 0
    Tuesday = 1
    Wednesday = 2
    Thursday = 3
    Friday = 4

class Lesson():
    def __init__(self, year, subject, building, room):
        self.Year = year
        self.Subject = subject
        self.Building = building
        self.Room = room

class WriteBlock():
    def __init__(self, sheetIndex, sheetName, rooms, startRow, endRow, startColumn, endColumn):
        self.SheetIndex = sheetIndex
        self.SheetName = sheetName
        self.Rooms = rooms
        self.StartRow = startRow
        self.EndRow = endRow
        self.StartColumn = startColumn
        self.EndColumn = endColumn


#### Function definitions ####
def extractClass(thisClass):
    regex = "^([0-9]{1,2})(.*)/([A-Z][a-z])[0-9]?( ([ABCDF])([0-9]{3}))?" # Block E not needed
    x = re.fullmatch(regex, thisClass)
    if (x):
        thisYear = x.group(1)
        thisSubject = x.group(3)
        thisBuilding = x.group(5)
        thisRoom = x.group(6)
        thisLesson = Lesson(thisYear, thisSubject, thisBuilding, thisRoom)
        return thisLesson
    else:
        return None

def getColumn(week, day, period):
    if (week == 2 and Days(day) == Days.Wednesday):
        day += 1
        return (day * 5) + period - 1 + day
    elif (week == 1):
        if (day > 2):
            day += 1
        return (day * 5) + period - 1 + day


#### Configuration of sheet to write to ####
# WriteBlock(sheetIndex, sheetName, rooms, startRow, endRow, startColumn, endColumn)

# KS4 Outside
KS4Outside0Block = WriteBlock(0, "KS4 outside", [""], 3, 8, 2, 37)
KS4Outside1Block = WriteBlock(0, "KS4 outside", [""], 12, 17, 2, 37)
# A Block
A0Rooms = ["A011","A012","A032","A034","A035","A038","A040","A041"]
A1Rooms = ["A105","A106","A110","A111","A113","A114","A117","A118","A119","A131"]
A0Block = WriteBlock(1, "A", A0Rooms, 3, 13, 2, 37)
A1Block = WriteBlock(1, "A", A1Rooms, 17, 27, 2, 37)
# B Block
B0Rooms = ["B015","B016","B019","B020","B023","B029","B034","B036","B039","B040","B043","B044","B047"]
B1Rooms = ["B113","B114","B117","B118","B121","B122","B132","B135","B139","B140","B143","B144"]
B0Block = WriteBlock(2, "B", B0Rooms, 3, 13, 2, 37)
B1Block = WriteBlock(2, "B", B1Rooms, 17, 27, 2, 37)
# C Block
C0Rooms = ["C007","C008","C009","C018","C019","C020","C021"]
C1Rooms = ["C103","C104","C105","C107","C108","C109","C110","C113","C114"]
C0Block = WriteBlock(3, "C", C0Rooms, 3, 13, 2, 37)
C1Block = WriteBlock(3, "C", C1Rooms, 17, 27, 2, 37)
# D Block
D0Rooms = ["D011","D013","D014","D016"]
D1Rooms = ["D005","D006","D007","D008","D026","D027","D027"]
D0Block = WriteBlock(4, "D", D0Rooms, 3, 13, 2, 37)
D1Block = WriteBlock(4, "D", D1Rooms, 17, 27, 2, 37)
# E Block not needed
# F Block
F0Rooms = ["F001","F002","F003","F004","F005","F006"] # F006 has been added as a dummy room for SMA
F1Rooms = ["F101","F102","F103","F104","F105","F106","F107"]
F0Block = WriteBlock(5, "F", F0Rooms, 3, 13, 2, 37)
F1Block = WriteBlock(5, "F", F1Rooms, 17, 27, 2, 37)
# Rooms
allRooms = {
    "A0" : A0Block,
    "A1" : A1Block,
    "B0" : B0Block,
    "B1" : B1Block,
    "C0" : C0Block,
    "C1" : C1Block,
    "F0" : F0Block,
    "F1" : F1Block
}

#### Initial communication with Excel sheets ####
# Excel timetable to read from
readLocation = "Whole staff timetable (as at 27th July).xlsx" # Give the location/name of the file
readWb = load_workbook(readLocation)
readSheet = readWb["rpttemp20200727152450"]
# Excel file to write to
writeLocation = "Supervision staffing by block v1 COVID.xlsx" # Give the location/name of the file
writeWb = load_workbook(writeLocation)

#### Configuration ####
# Rows of timetable being read (starting at first teacher)
startRow = 6
endRow = 114
# Columns of timetable being read (starting at name of teacher)
startColumn = 1
endColumn = 53
# Timetable details
lessonsPerDay = 5
numOfWeeks = 2
daysPerWeek = 5

codes = ["", "DUTY", "Fac. time/Cover", "HoY duties", "Mentor", "NQT time", "NTE", "PPA", "SLT Time", "SciTT", "TLR Time"] # Codes to be considered available
exclStaff = ["AMR", "GKA", "KJA"] # Staff to exclude

#### Notes ####
# SciTT (except if it is on a Friday)
# DUTY (P2 only - not P4) (Break time only, not lunch time)
# SMA has been given room F006 (non-existent) to indicate downstairs F-Block as per an email from Matt
# Deleted column Ex:6 from each day (10 total) to match up 5 periods easily
###############

for r in range(startRow, endRow):
    newWeek = []
    for c in range(startColumn, endColumn):
        thisCell = readSheet.cell(r, c).value
        if (c == 1):
            staffName = thisCell
        elif (c == 2):
            staffCode = thisCell
        else:
            if (thisCell):
                newWeek.append(thisCell)
            else:
                newWeek.append("")
    print("--------------------")
    print(staffCode, staffName)
    print("--------------------")
    if (staffCode not in exclStaff):
        for i in range(numOfWeeks * lessonsPerDay * daysPerWeek): # Iterate through the current 1 week of classes for a member of staff
            thisClass = newWeek[i][:-1]
            thisLesson = extractClass(thisClass)
            thisWeek = (i // ((lessonsPerDay) * daysPerWeek)) + 1
            thisDayNum = ((i // lessonsPerDay) % daysPerWeek)
            thisDay = Days(thisDayNum)
            thisPeriod = (i % (lessonsPerDay)) + 1
            if (thisWeek == 1 or thisDay == Days.Wednesday):
                # Could negate expressions but left this way round for ease of readability
                if (thisClass == "SciTT" and thisDay == Days.Friday):
                    # Ignore SciTT code if it is on a Friday
                    pass
                elif (thisClass == "DUTY" and thisPeriod != 2):
                    # Ignore DUTY code when not P2
                    pass
                else:
                    if (thisPeriod != 1):
                        prevClass = newWeek[i-1][:-1] # Cell to the left
                        ready = False
                        if (thisClass in codes and prevClass in codes):
                            # Two available blocks adjacent to each other
                            blocks = [KS4Outside0Block, KS4Outside1Block] # KS4 outside
                            ready = True
                        else:
                            thisLesson = extractClass(thisClass)
                            prevLesson = extractClass(prevClass)
                            fullRoom = None
                            if (thisClass in codes and prevLesson):
                                # Current is an available code and left cell is a lesson e.g. 11MAR/Ma D007
                                fullRoom = prevLesson.Building + prevLesson.Room # e.g. A113
                            elif (thisLesson and prevClass in codes):
                                # Current is a lesson e.g. 11MAR/Ma D007 and left cell is an available code
                                fullRoom = thisLesson.Building + thisLesson.Room # e.g. A113
                            if (fullRoom):
                                buildingFloor = fullRoom[:2]
                                if (buildingFloor == "D0"):
                                    for b in [D0Block, D1Block]:
                                        if (fullRoom in b.Rooms):
                                            blocks = [b]
                                            ready = True
                                else:
                                    blocks = [allRooms[buildingFloor]] # e.g. A1 (key to allRooms dictionary) to access A1Block
                                    if (fullRoom in blocks[0].Rooms):
                                        # Now identified: one is a lesson and one is an available code AND the lesson is included in list of room codes for that block
                                        ready = True
                        if (ready):
                            # Now ready to copy 3 letter code into slot:
                            for b in blocks:
                                thisBlock = b
                                writeWs = writeWb[thisBlock.SheetName]
                                writeColumn = getColumn(thisWeek, thisDayNum, thisPeriod)
                                if (writeColumn):
                                    writeColumn += thisBlock.StartColumn
                                    for r in range(thisBlock.StartRow, thisBlock.EndRow + 1):
                                        if (writeWs.cell(r, writeColumn).value == None):
                                            print(staffCode + " to (" + str(r) + ", " + str(writeColumn) + ")")
                                            writeWs.cell(r, writeColumn).value = staffCode
                                            writeWb.save(writeLocation)
                                            break